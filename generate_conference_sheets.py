"""
generate_conference_sheets.py
Warmth — GTM Hackathon demo script

Simulates one day at the GTM Hackathon:
  1. Pre-meet: who to target today (from attendee roster + pre-meet drafts)
  2. Live capture: 5 conversations, each run through the meet pipeline
  3. Post-meet: score every person met, generate follow-up drafts

Outputs a single Excel workbook: warmth_conference_demo.xlsx
  Sheet 1 — Pre-Meet Targets  (who to meet, why, enrichment)
  Sheet 2 — Conversations Log  (live capture — transcript excerpt, person context)
  Sheet 3 — Connections Scored (warmth score, ICP, routing, narrative)
  Sheet 4 — Follow-up Drafts  (Gmail-ready subject + body for each hot/warm lead)

Usage (from gtmhackathon/):
  warmth/.venv/bin/python generate_conference_sheets.py
"""

import json
import os
import sys
import glob
import re
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Run: warmth/.venv/bin/python -m pip install openpyxl")

# ─── paths ───────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
WARMTH_DRAFTS = ROOT / "warmth" / "drafts"
OUTER_DRAFTS  = ROOT / "drafts"
ATTENDEES_JSON = ROOT / "warmth" / "data" / "gtm_hackathon_attendees.json"
OUT_FILE = ROOT / "warmth_conference_demo.xlsx"

# ─── colour palette (Warmth brand) ───────────────────────────────────────────
C_RED      = "DC2626"   # hot / ICP match
C_EMBER    = "C2410C"   # primary action
C_ORANGE   = "EA580C"   # flame / warm band
C_AMBER    = "F59E0B"   # pending / medium
C_STONE    = "A8A29E"   # cold
C_INK      = "160D07"   # near-black text
C_INK_MED  = "57534E"   # secondary text
C_SURFACE  = "FFF8F3"   # warm white bg
C_HEADER   = "1C0A03"   # dark header bg
C_ACCENT   = "FEDDCA"   # light peach for alt rows


# ─── styling helpers ─────────────────────────────────────────────────────────
def _side():
    return Side(style="thin", color="E7D9D0")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill():
    return PatternFill("solid", fgColor=C_HEADER)

def _alt_fill():
    return PatternFill("solid", fgColor=C_ACCENT)

def _band_fill(band: str) -> PatternFill:
    colour = {"hot": C_RED, "warm": C_ORANGE, "cold": C_STONE}.get(band.lower(), C_STONE)
    return PatternFill("solid", fgColor=colour)

def _header_font():
    return Font(name="Calibri", bold=True, color="FFFFFF", size=10)

def _body_font(bold=False, color=C_INK):
    return Font(name="Calibri", bold=bold, color=color, size=9)

def _wrap():
    return Alignment(wrap_text=True, vertical="top")

def _centre():
    return Alignment(horizontal="center", vertical="center")

def _apply_header_row(ws, row_num, cols):
    """Style a header row."""
    for col_idx, label in enumerate(cols, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=label)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _centre()
        cell.border = _border()

def _apply_data_row(ws, row_num, values, alt=False, bold=False, band=None):
    fill = _band_fill(band) if band else (_alt_fill() if alt else PatternFill("solid", fgColor="FFFFFF"))
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = _body_font(bold=bold)
        cell.fill = fill
        cell.alignment = _wrap()
        cell.border = _border()

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def _title_row(ws, title, n_cols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", bold=True, size=13, color=C_EMBER)
    cell.fill = PatternFill("solid", fgColor="FEF3E8")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26


# ─── data loading ─────────────────────────────────────────────────────────────
def load_attendees():
    if ATTENDEES_JSON.exists():
        return json.loads(ATTENDEES_JSON.read_text())
    return []


def load_premeet_drafts():
    """Parse all pre_meet_intro drafts into structured dicts."""
    people = {}
    for folder in [WARMTH_DRAFTS, OUTER_DRAFTS]:
        if not folder.exists():
            continue
        for f in sorted(folder.glob("*.json")):
            try:
                d = json.loads(f.read_text())
            except Exception:
                continue
            if d.get("purpose") != "pre_meet_intro":
                continue
            subject = d.get("subject", "")
            name = subject.replace("Excited to connect at the conference, ", "").strip()
            if not name or name in people:
                continue
            body = d.get("body", "")
            company = ""
            interests = ""
            m = re.search(r"your work at (.+?)\.", body)
            if m:
                company = m.group(1).strip()
            for line in body.split("\n"):
                if "interests:" in line and "CONTEXT" not in line:
                    interests = line.split("interests:")[1].strip()
                    break
            people[name] = {
                "name": name,
                "company": company,
                "interests": interests,
                "gmail_url": d.get("gmail_compose_url", ""),
            }
    return list(people.values())


def load_postmeet_drafts():
    """Parse all post_meet_followup drafts into structured dicts."""
    people = {}
    for folder in [WARMTH_DRAFTS, OUTER_DRAFTS]:
        if not folder.exists():
            continue
        for f in sorted(folder.glob("*.json")):
            try:
                d = json.loads(f.read_text())
            except Exception:
                continue
            if d.get("purpose") != "post_meet_followup":
                continue
            subject = d.get("subject", "")
            name = subject.replace("Great meeting you, ", "").strip()
            if not name or name in people:
                continue
            body = d.get("body", "")
            # parse scores block
            scores = {}
            person = {}
            for line in body.split("\n"):
                for k in ["icp_score", "warmth_score", "predicted_score",
                          "actual_score", "band", "uplift", "routing"]:
                    if f"  {k}:" in line:
                        scores[k] = line.split(f"  {k}:")[1].strip()
                for k in ["summary", "communication_style", "values",
                          "dominant_topic", "learnings", "pain_points"]:
                    if f"  {k}:" in line:
                        person[k] = line.split(f"  {k}:")[1].strip()
            # extract lead company
            company = ""
            m = re.search(r"ideas for (.+?)\.", body)
            if m:
                company = m.group(1).strip()
            people[name] = {
                "name": name,
                "company": company,
                "to": d.get("to") or "",
                "scores": scores,
                "person": person,
                "subject": d.get("subject", ""),
                "body": body,
                "gmail_url": d.get("gmail_compose_url", ""),
            }
    return list(people.values())


# ─── simulated conversation data ─────────────────────────────────────────────
# Rich fictional conversations that look like what the Warmth pipeline actually
# produces — modelled directly on the real draft output format.
CONVERSATIONS = [
    {
        "name": "Anna",
        "company": "Acme RevOps",
        "time": "09:42",
        "duration_s": 95,
        "trigger": "phrase",
        "transcript_excerpt": (
            "I run RevOps at Acme. We basically have three reps who keep going "
            "off-piste in HubSpot. The manual data entry is honestly killing us — "
            "like we spend more time cleaning the CRM than actually selling. "
            "I just found out HubSpot has AI forecasting now which is interesting "
            "but honestly I care way more about accuracy than features."
        ),
        "icp_score": 82,
        "warmth_score": 71,
        "predicted_score": 60,
        "actual_score": 71,
        "band": "hot",
        "uplift": 11,
        "routing": "crm_and_outreach",
        "communication_style": "analytical, data-driven, relational",
        "values": "accuracy",
        "dominant_topic": "run revops",
        "learnings": "HubSpot has AI forecasting",
        "pain_points": "manual data entry across the team (moderate)",
        "narrative": (
            "Anna is analytical, data-driven, and relational — cares deeply about "
            "accuracy. Dominant topic: RevOps process. Recently learned HubSpot has "
            "AI forecasting. Moderate-to-high pain around manual data entry."
        ),
    },
    {
        "name": "Federico Ruosi",
        "company": "Atomico",
        "time": "10:15",
        "duration_s": 72,
        "trigger": "tap",
        "transcript_excerpt": (
            "We're Series A–Growth at Atomico, mainly looking at deep tech plays, "
            "industrial automation, internet infrastructure. Good connections are "
            "the real currency here honestly. Warm intros move faster than cold "
            "decks by like 10x."
        ),
        "icp_score": 45,
        "warmth_score": 68,
        "predicted_score": 50,
        "actual_score": 68,
        "band": "hot",
        "uplift": 18,
        "routing": "crm_and_outreach",
        "communication_style": "visionary, relational",
        "values": "quality networks, warm intros",
        "dominant_topic": "deep tech investing",
        "learnings": "warm intros move 10x faster than cold decks",
        "pain_points": "cold outreach from founders (low)",
        "narrative": (
            "Federico is visionary and relational — network quality over volume. "
            "Dominant topic: deep tech. Key insight: warm intros 10x cold decks. "
            "Potential intro path for portfolio founders."
        ),
    },
    {
        "name": "Sam Rivera",
        "company": "Glide",
        "time": "11:03",
        "duration_s": 58,
        "trigger": "phrase",
        "transcript_excerpt": (
            "We're rebuilding RevOps basically from scratch. The old Salesforce "
            "setup was a nightmare. We're actually considering just going full "
            "HubSpot. I want something my team can actually use without training."
        ),
        "icp_score": 70,
        "warmth_score": 42,
        "predicted_score": 55,
        "actual_score": 42,
        "band": "cold",
        "uplift": -13,
        "routing": "founder_community",
        "communication_style": "pragmatic",
        "values": "simplicity, team adoption",
        "dominant_topic": "rebuilding revops",
        "learnings": "migrating from Salesforce to HubSpot",
        "pain_points": "tool complexity, adoption (moderate)",
        "narrative": (
            "Sam is pragmatic — wants simplicity above features. Rebuilding RevOps, "
            "evaluating HubSpot. Warmth flat vs. prediction; route to founder "
            "community for a warmer intro path."
        ),
    },
    {
        "name": "Reem Wyndham",
        "company": "Pact VC",
        "time": "13:22",
        "duration_s": 84,
        "trigger": "tap",
        "transcript_excerpt": (
            "Pact focuses on pre-seed and seed fintech and SaaS. We like seeing "
            "GTM traction before we write a check. Revenue signals, not just "
            "pipeline. If you've got your first 5 customers that's the conversation "
            "I want to have."
        ),
        "icp_score": 38,
        "warmth_score": 74,
        "predicted_score": 55,
        "actual_score": 74,
        "band": "hot",
        "uplift": 19,
        "routing": "crm_and_outreach",
        "communication_style": "data-driven, skeptical",
        "values": "traction signals, revenue proof",
        "dominant_topic": "pre-seed fintech investing",
        "learnings": "first 5 customers unlock the real conversation",
        "pain_points": "founders pitching without traction (moderate)",
        "narrative": (
            "Reem is data-driven and mildly skeptical — wants revenue signals not "
            "pipeline. Pre-seed/seed fintech & SaaS focus. Warmth spiked: high "
            "intent conversation. Crux: 'first 5 customers'."
        ),
    },
    {
        "name": "Nick Stocks",
        "company": "White Star Capital",
        "time": "15:48",
        "duration_s": 66,
        "trigger": "phrase",
        "transcript_excerpt": (
            "Series A digital platforms and fintech is our sweet spot. We've done "
            "a lot of work on GTM motion — like how companies go from product-market "
            "fit to repeatable revenue. That's the gap most startups fall into."
        ),
        "icp_score": 40,
        "warmth_score": 65,
        "predicted_score": 52,
        "actual_score": 65,
        "band": "warm",
        "uplift": 13,
        "routing": "crm_and_outreach",
        "communication_style": "analytical, visionary",
        "values": "repeatability, GTM execution",
        "dominant_topic": "PMF to repeatable revenue",
        "learnings": "gap between PMF and repeatable revenue is where startups fall",
        "pain_points": "inconsistent GTM motion in portfolio (moderate)",
        "narrative": (
            "Nick is analytical + visionary — passionate about the PMF→revenue gap. "
            "Series A fintech/digital platforms. Warmth above prediction; genuine "
            "interest in GTM tools that solve the repeatability problem."
        ),
    },
]


# ─── sheet builders ───────────────────────────────────────────────────────────

def build_premeet_sheet(wb, premeet_people, attendees):
    ws = wb.create_sheet("1. Pre-Meet Targets")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    n_cols = 7
    _title_row(ws, "Warmth — Pre-Meet Target List · GTM Hackathon, 20 June 2026", n_cols, row=1)

    cols = ["Name", "Company", "Interests / Signal", "LinkedIn", "Warmth Priority",
            "Pre-Meet Outreach?", "Notes"]
    _apply_header_row(ws, 2, cols)

    # Merge attendee email into premeet people where possible
    att_map = {a["name"]: a for a in attendees}

    # Sort: attendees first (we have richer data), then scraped targets
    all_targets = list(premeet_people)
    for a in attendees:
        if a["name"] not in [p["name"] for p in premeet_people]:
            all_targets.append({
                "name": a["name"],
                "company": a.get("company") or "—",
                "interests": ", ".join(a.get("interests", [])),
                "gmail_url": "",
            })

    priority_map = {
        "Federico Ruosi": "High — investor (Series A–Growth)",
        "Nick Stocks": "High — Series A, GTM aligned",
        "Aaron Sweet": "High — Vantage Capital, fintech",
        "Reem Wyndham": "High — Pact VC, fintech/SaaS",
        "Boris Dorin": "Med — Deep Tech AI",
        "Ella Mamelok": "Med — Dalloway Partners",
        "Liz Actub": "Med — Osun Capital fintech",
        "Kieren Auluk": "Med — early stage general",
    }

    for row_idx, person in enumerate(all_targets, 3):
        name = person["name"]
        att = att_map.get(name, {})
        linkedin = att.get("linkedin", "")
        has_draft = "Yes — draft ready" if person.get("gmail_url") else "No"
        priority = priority_map.get(name, "Standard")
        notes = att.get("research_notes", "")[:120].replace("\n", " ") if att.get("research_notes") else ""
        alt = (row_idx % 2 == 0)
        _apply_data_row(ws, row_idx,
            [name, person["company"], person["interests"][:80], linkedin,
             priority, has_draft, notes],
            alt=alt)
        ws.row_dimensions[row_idx].height = 30

    _set_col_widths(ws, [22, 26, 44, 40, 30, 20, 55])
    _freeze(ws, "A3")
    return ws


def build_conversations_sheet(wb):
    ws = wb.create_sheet("2. Live Conversations")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    n_cols = 9
    _title_row(ws, "Warmth — Live Capture Log · GTM Hackathon, 20 June 2026", n_cols, row=1)

    cols = ["Time", "Name", "Company", "Trigger", "Duration",
            "Transcript Excerpt", "Topic", "Learnings", "Pain Points"]
    _apply_header_row(ws, 2, cols)

    for row_idx, c in enumerate(CONVERSATIONS, 3):
        dur = f"{c['duration_s']}s"
        trig = "Wake phrase" if c["trigger"] == "phrase" else "Manual tap"
        alt = (row_idx % 2 == 0)
        _apply_data_row(ws, row_idx, [
            c["time"], c["name"], c["company"], trig, dur,
            c["transcript_excerpt"], c["dominant_topic"],
            c["learnings"], c["pain_points"],
        ], alt=alt)
        ws.row_dimensions[row_idx].height = 52

    _set_col_widths(ws, [8, 18, 22, 16, 10, 60, 26, 38, 36])
    _freeze(ws, "A3")
    return ws


def build_scored_sheet(wb):
    ws = wb.create_sheet("3. Connections Scored")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    n_cols = 10
    _title_row(ws, "Warmth — Post-Meet Scoring · GTM Hackathon, 20 June 2026", n_cols, row=1)

    cols = ["Name", "Company", "ICP Score", "Warmth Score",
            "Predicted", "Actual", "Uplift", "Band", "Routing", "Person Narrative"]
    _apply_header_row(ws, 2, cols)

    for row_idx, c in enumerate(CONVERSATIONS, 3):
        uplift_str = f"+{c['uplift']}" if c["uplift"] >= 0 else str(c["uplift"])
        routing_label = (
            "→ CRM + outreach" if c["routing"] == "crm_and_outreach"
            else "→ Founder community"
        )
        band = c["band"]
        _apply_data_row(ws, row_idx, [
            c["name"], c["company"],
            c["icp_score"], c["warmth_score"],
            c["predicted_score"], c["actual_score"],
            uplift_str, band.upper(), routing_label,
            c["narrative"],
        ], band=band)
        ws.row_dimensions[row_idx].height = 52

        # Colour the Band cell white text so it's readable on the coloured bg
        band_cell = ws.cell(row=row_idx, column=8)
        band_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        band_cell.alignment = _centre()

    _set_col_widths(ws, [18, 22, 12, 14, 12, 12, 10, 10, 22, 58])
    _freeze(ws, "A3")
    return ws


def build_followup_sheet(wb):
    ws = wb.create_sheet("4. Follow-up Drafts")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    n_cols = 7
    _title_row(ws, "Warmth — Gmail Follow-up Drafts (Lightfern handoff) · GTM Hackathon, 20 June 2026",
               n_cols, row=1)

    cols = ["Name", "Company", "Band", "Subject Line", "Email Body (draft)",
            "Communication Style", "Key Hook (person context)"]
    _apply_header_row(ws, 2, cols)

    # Only include leads routed to CRM / outreach, sorted hot first
    routed = [c for c in CONVERSATIONS if c["routing"] == "crm_and_outreach"]
    routed.sort(key=lambda x: x["actual_score"], reverse=True)

    email_bodies = {
        "Anna": (
            "Hi Anna,\n\n"
            "Really enjoyed our chat about RevOps at Acme — the manual data entry "
            "problem you described is exactly the kind of signal Warmth is built to "
            "catch before it kills a pipeline.\n\n"
            "Worth 20 mins to show you how we're approaching the accuracy problem? "
            "Happy to go through the HubSpot AI forecasting angle too given what "
            "you mentioned.\n\n"
            "[Lightfern will polish this before send]"
        ),
        "Federico Ruosi": (
            "Hi Federico,\n\n"
            "Great to meet you at the GTM Hackathon — your point about warm intros "
            "moving 10x faster than cold decks landed. That's literally the problem "
            "Warmth solves: we instrument the in-person moment so the follow-up "
            "feels warm because it IS warm.\n\n"
            "Worth a quick call to explore whether Warmth fits any of your Atomico "
            "portfolio's GTM stack?\n\n"
            "[Lightfern will polish this before send]"
        ),
        "Reem Wyndham": (
            "Hi Reem,\n\n"
            "Loved the framing — 'first 5 customers unlock the real conversation.' "
            "We're building exactly the infrastructure that surfaces those early "
            "traction signals from in-person events before they evaporate.\n\n"
            "Would love to show you the demo and hear whether Pact sees this "
            "fitting the fintech/SaaS founders you back.\n\n"
            "[Lightfern will polish this before send]"
        ),
        "Nick Stocks": (
            "Hi Nick,\n\n"
            "Your framing of the PMF→repeatable revenue gap stuck with me. Warmth "
            "lives in that gap — it captures the relationship signal at conferences "
            "that usually gets lost and routes it directly into CRM + outreach.\n\n"
            "Happy to walk through how White Star portfolio companies could use "
            "this for their Series A GTM motion.\n\n"
            "[Lightfern will polish this before send]"
        ),
    }

    hooks = {
        "Anna": "Pain: manual data entry. Learning: HubSpot AI forecasting. Values: accuracy.",
        "Federico Ruosi": "Warm intros 10× cold decks. Portfolio founder intro potential.",
        "Reem Wyndham": "'First 5 customers unlock the real conversation' — key phrase.",
        "Nick Stocks": "PMF→repeatable revenue gap. GTM execution focus.",
    }

    for row_idx, c in enumerate(routed, 3):
        name = c["name"]
        subject = f"Great meeting you at the GTM Hackathon, {name.split()[0]}"
        body = email_bodies.get(name, f"Hi {name.split()[0]},\n\nGreat to meet you. [Lightfern will polish]")
        hook = hooks.get(name, c["narrative"][:100])
        alt = (row_idx % 2 == 0)
        _apply_data_row(ws, row_idx, [
            name, c["company"], c["band"].upper(),
            subject, body, c["communication_style"], hook,
        ], alt=alt)
        ws.row_dimensions[row_idx].height = 72

        # Colour band cell
        band_cell = ws.cell(row=row_idx, column=3)
        band_cell.fill = _band_fill(c["band"])
        band_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        band_cell.alignment = _centre()

    _set_col_widths(ws, [18, 22, 10, 44, 64, 30, 48])
    _freeze(ws, "A3")
    return ws


def build_summary_sheet(wb):
    """Dashboard-style first sheet with key stats."""
    ws = wb.create_sheet("0. Conference Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    cell = ws.cell(row=1, column=1,
                   value="Warmth — Conference Intelligence · GTM Hackathon, 20 June 2026")
    cell.font = Font(name="Calibri", bold=True, size=16, color=C_EMBER)
    cell.fill = PatternFill("solid", fgColor="FEF3E8")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    sub = ws.cell(row=2, column=1,
                  value="Best GTM is real connection. Warmth makes every conversation count.")
    sub.font = Font(name="Calibri", italic=True, size=11, color=C_INK_MED)
    sub.fill = PatternFill("solid", fgColor="FEF3E8")
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10  # spacer

    # ── stat cards ──────────────────────────────────────────────────────────
    stats = [
        ("Attendees Researched", "10", C_INK,     "Pre-meet targets identified"),
        ("Conversations Captured", "5",  C_ORANGE, "Live today via Warmth"),
        ("Hot Leads (uplift)",    "3",  C_RED,    "Warmth rose vs. prediction"),
        ("Warm Leads",            "1",  C_AMBER,  "Above threshold, monitor"),
        ("Cold / Community",      "1",  C_STONE,  "Routed to founder network"),
        ("Follow-up Drafts",      "4",  C_EMBER,  "Gmail-ready via Lightfern"),
    ]

    stat_col_start = 1
    for i, (label, value, colour, sub_label) in enumerate(stats):
        col = stat_col_start + i
        # Value
        vc = ws.cell(row=4, column=col, value=value)
        vc.font = Font(name="Calibri", bold=True, size=24, color=colour)
        vc.fill = PatternFill("solid", fgColor="FFF8F3")
        vc.alignment = _centre()
        ws.row_dimensions[4].height = 36

        # Label
        lc = ws.cell(row=5, column=col, value=label)
        lc.font = Font(name="Calibri", bold=True, size=9, color=C_INK)
        lc.fill = PatternFill("solid", fgColor="FFF8F3")
        lc.alignment = _centre()
        ws.row_dimensions[5].height = 16

        # Sub-label
        sc = ws.cell(row=6, column=col, value=sub_label)
        sc.font = Font(name="Calibri", italic=True, size=8, color=C_INK_MED)
        sc.fill = PatternFill("solid", fgColor="FFF8F3")
        sc.alignment = _centre()
        ws.row_dimensions[6].height = 14

    ws.row_dimensions[7].height = 10  # spacer

    # ── timeline ────────────────────────────────────────────────────────────
    ws.merge_cells("A8:H8")
    tl_hdr = ws.cell(row=8, column=1, value="Day Timeline")
    tl_hdr.font = Font(name="Calibri", bold=True, size=11, color=C_EMBER)
    tl_hdr.fill = PatternFill("solid", fgColor="FEF3E8")
    tl_hdr.alignment = Alignment(horizontal="left", indent=1, vertical="center")
    ws.row_dimensions[8].height = 20

    timeline_cols = ["Time", "Person", "Company", "Band", "Uplift", "Key Signal", "Routed To"]
    _apply_header_row(ws, 9, timeline_cols[:7])
    ws.row_dimensions[9].height = 18

    for row_idx, c in enumerate(CONVERSATIONS, 10):
        uplift_str = f"+{c['uplift']}" if c["uplift"] >= 0 else str(c["uplift"])
        routing_label = "CRM + outreach" if c["routing"] == "crm_and_outreach" else "Founder community"
        _apply_data_row(ws, row_idx, [
            c["time"], c["name"], c["company"],
            c["band"].upper(), uplift_str,
            c["learnings"][:60], routing_label,
        ], band=c["band"])
        ws.row_dimensions[row_idx].height = 22

        # White text on coloured band cell
        band_cell = ws.cell(row=row_idx, column=4)
        band_cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        band_cell.alignment = _centre()

    ws.row_dimensions[15].height = 10  # spacer after timeline

    # ── routing explanation ──────────────────────────────────────────────────
    ws.merge_cells("A16:H16")
    exp_hdr = ws.cell(row=16, column=1, value="How Warmth Routes Connections")
    exp_hdr.font = Font(name="Calibri", bold=True, size=11, color=C_EMBER)
    exp_hdr.fill = PatternFill("solid", fgColor="FEF3E8")
    exp_hdr.alignment = Alignment(horizontal="left", indent=1, vertical="center")
    ws.row_dimensions[16].height = 20

    routing_rows = [
        ("HOT (uplift > 0, score ≥ 70)", C_RED,
         "Warmth rose vs. pre-meet prediction",
         "→ Zero CRM push + Lightfern Gmail draft + Faxxing outreach sequence"),
        ("WARM (uplift > 0, score 50–69)", C_ORANGE,
         "Warmth up, relationship building",
         "→ CRM + lighter outreach touch"),
        ("COLD (uplift ≤ 0 or score < 50)", C_STONE,
         "Warmth flat or fell vs. prediction",
         "→ Routed to founder community (nearest-neighbour match)"),
    ]

    routing_header = ["Band", "Condition", "Meaning", "Action"]
    _apply_header_row(ws, 17, routing_header[:4])
    ws.row_dimensions[17].height = 18

    for r_idx, (band_label, colour, condition, action) in enumerate(routing_rows, 18):
        fill = PatternFill("solid", fgColor=colour)
        alt_fill = PatternFill("solid", fgColor="FFF8F3")
        for col_idx, val in enumerate([band_label, condition, "", action], 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.fill = fill if col_idx == 1 else alt_fill
            cell.font = Font(name="Calibri",
                             bold=(col_idx == 1),
                             color="FFFFFF" if col_idx == 1 else C_INK,
                             size=9)
            cell.alignment = _wrap()
            cell.border = _border()
        ws.cell(row=r_idx, column=3, value=condition).fill = alt_fill
        ws.cell(row=r_idx, column=3).font = Font(name="Calibri", italic=True, color=C_INK_MED, size=9)
        ws.row_dimensions[r_idx].height = 20

    _set_col_widths(ws, [22, 22, 30, 20, 12, 48, 24, 10])
    return ws


# ─── main ─────────────────────────────────────────────────────────────────────
def main():
    print("Warmth Conference Demo Sheet Generator")
    print("=" * 40)

    premeet_people = load_premeet_drafts()
    attendees = load_attendees()
    postmeet_people = load_postmeet_drafts()

    print(f"  Loaded {len(premeet_people)} pre-meet targets from drafts")
    print(f"  Loaded {len(attendees)} attendees from calendar data")
    print(f"  Loaded {len(postmeet_people)} post-meet records")
    print(f"  {len(CONVERSATIONS)} simulated conversations")
    print()

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building sheets...")
    build_summary_sheet(wb)
    print("  ✓  Sheet 0: Conference Summary")
    build_premeet_sheet(wb, premeet_people, attendees)
    print("  ✓  Sheet 1: Pre-Meet Targets")
    build_conversations_sheet(wb)
    print("  ✓  Sheet 2: Live Conversations")
    build_scored_sheet(wb)
    print("  ✓  Sheet 3: Connections Scored")
    build_followup_sheet(wb)
    print("  ✓  Sheet 4: Follow-up Drafts")

    wb.save(OUT_FILE)
    print()
    print(f"Saved → {OUT_FILE}")
    print()
    print("Demo narrative:")
    print("  Morning: 10 targets researched, pre-meet emails drafted via Lightfern")
    print("  On the floor: 5 conversations captured (3 phrase trigger, 2 manual tap)")
    print("  ML pipeline: warmth scored vs. pre-meet prediction → uplift computed")
    print("  Evening: 3 HOT, 1 WARM → CRM push + Gmail draft; 1 COLD → founder community")
    print()
    print("Best GTM is real connection. Warmth makes every conversation count.")


if __name__ == "__main__":
    main()
